import time
import requests
import pandas as pd
import re, unicodedata
from collections import OrderedDict
from openpyxl.utils import get_column_letter
from datetime import datetime

# Settings
USER_AGENT = "Jimmy-DataPipeline/1.0 (ttnkashe@mtu.edu)"
PAUSE_SECONDS = 0.25
SELECT_FORM = "10-K"  # '10-K' | '10-Q' | 'latest'
OUT_FILE = "SEC_BS_IS_SP_500.xlsx"
START_YEAR = 2020  # only use filings from 2020 or newer
END_YEAR = 2025    # extract up to 2025
STRICT_YEAR = False  # True = drop older values; False = fallback to oldest available

HEADERS = {"User-Agent": USER_AGENT, "Accept-Encoding": "gzip, deflate"}

# Updated with all requested companies
REQUESTED_NAMES = [
    "Microsoft Corp", "Nvidia Corp", "Apple Inc.", "Amazon.com Inc", "Meta Platforms, Inc. Class A",
    "Broadcom Inc.", "Alphabet Inc. Class A", "Tesla, Inc.", "Berkshire Hathaway Class B",
    "Alphabet Inc. Class C", "Jpmorgan Chase & Co.", "Visa Inc.", "Eli Lilly & Co.", "Netflix Inc",
    "Mastercard Incorporated", "Costco Wholesale Corp", "Exxon Mobil Corporation", "Walmart Inc.",
    "Procter & Gamble Company", "Johnson & Johnson", "Home Depot, Inc.", "Abbvie Inc.",
    "Bank of America Corporation", "Unitedhealth Group Incorporated", "Coca-Cola Company",
    "Philip Morris International Inc.", "Salesforce, Inc.", "Oracle Corp", "Cisco Systems, Inc.",
    "Ge Aerospace", "Palantir Technologies Inc. Class A", "International Business Machines Corporation",
    "Wells Fargo & Co.", "Abbott Laboratories", "Mcdonald's Corporation", "Chevron Corporation",
    "Linde Plc", "Servicenow, Inc.", "The Walt Disney Company", "Accenture Plc", "At&t Inc.",
    "Intuitive Surgical Inc.", "Merck & Co., Inc.", "Uber Technologies, Inc.", "Goldman Sachs Group Inc.",
    "Intuit Inc", "Verizon Communications", "Advanced Micro Devices", "Adobe Inc.", "Rtx Corporation",
    "Pepsico, Inc.", "Booking Holdings Inc.", "Texas Instruments Incorporated", "Qualcomm Inc",
    "Progressive Corporation", "Caterpillar Inc.", "S&p Global Inc.", "American Express Company",
    "Morgan Stanley", "Boston Scientific Corp.", "Boeing Company", "Thermo Fisher Scientific, Inc.",
    "The Charles Schwab Corporation", "Tjx Companies, Inc.", "Nextra Energy, Inc.", "Amgen Inc",
    "Honeywell International, Inc.", "Blackrock, Inc.", "Citigroup Inc.", "Union Pacific Corp.",
    "Gilead Sciences Inc", "Comcast Corp", "Applied Materials Inc", "Automatic Data Processing",
    "Pfizer Inc.", "Stryker Corporation", "Deere & Company", "Lowe's Companies Inc.",
    "Eaton Corporation, Plcs", "Ge Vernova Inc.", "Palo Alto Networks, Inc.", "Danaher Corporation",
    "Capital One Financial", "T-Mobile Us, Inc.", "Marsh & Mclennan Companies, Inc.",
    "Vertex Pharmaceuticals Inc", "Conocophillips", "Analog Devices, Inc.", "Medtronic Plc",
    "Chubb Limited", "Crowdstrike Holdings, Inc. Class A", "Micron Technology, Inc.", "Lam Research Corp",
    "Amphenol Corporation", "Kla Corporation", "Cme Group Inc.", "Altria Group, Inc.", "Blackstone Inc.",
    "Intercontinental Exchange Inc.", "American Tower Corporation", "Lockheed Martin Corp.",
    "The Southern Company", "Prologis, Inc.", "Arista Networks", "Bristol-Myers Squibb Co.",
    "Trane Technologies Plc", "Starbucks Corp", "Elevance Health, Inc.", "Fiserv, Inc.",
    "Duke Energy Corporation", "Welltower Inc.", "Mckesson Corporation", "Constellation Energy Corporation",
    "Intel Corp", "Cadence Design Systems", "The Cigna Group", "Arthur J. Gallagher & Co.",
    "Waste Management, Inc.", "Parker-Hannifin Corporation", "Mondelez International, Inc. Class A",
    "Equinix, Inc. Reit", "The Sherwin-Williams Company", "3m Company", "Kkr & Co. Inc.",
    "Transdigm Group Incorporated", "O'reilly Automotive, Inc.", "Cvs Health Corporation", "Synopsys Inc",
    "Aon Plc Class A", "Cintas Corp", "Colgate-Palmolive Company", "Moody's Corporation", "Zoetis Inc.",
    "Motorola Solutions, Inc.", "Paypal Holdings, Inc.", "Nike, Inc.", "Williams Companies Inc.",
    "General Dynamics Corporation", "United Parcel Service, Inc. Class B", "Doordash, Inc. Class A",
    "Chipotle Mexican Grill, Inc.", "Hca Healthcare, Inc.", "Pnc Financial Services Group", "U.S. Bancorp",
    "Howmet Aerospace Inc.", "Ecolab, Inc.", "Emerson Electric Co.", "Illinois Tool Works Inc.",
    "Fortinet, Inc.", "Autozone, Inc.", "Northrop Grumman Corp.", "Johnson Controls International Plc",
    "Bank of New York Mellon Corporation", "Regeneron Pharmaceuticals Inc", "Autodesk Inc",
    "Eog Resources, Inc.", "The Travelers Companies, Inc.", "Roper Technologies, Inc.",
    "Air Products & Chemicals, Inc.", "Newmont Corporation", "Marriot International Class A",
    "Hilton Worldwide Holdings Inc.", "Royal Caribbean Group", "Csx Corporation",
    "Apollo Global Management, Inc.", "Carrier Global Corporation", "Workday, Inc. Class A",
    "Airbnb, Inc. Class A", "American Electric Power Company, Inc.", "Coinbase Global, Inc. Class A",
    "Freeport-Mcmoran Inc.", "The Allstate Corporation", "Norfolk Southern Corp.", "Copart Inc",
    "Axon Enterprise, Inc.", "Kinder Morgan, Inc.", "Digital Realty Trust, Inc.",
    "Nxp Semiconductors N.v.", "Vistra Corp.", "Truist Financial Corporation", "Aflac Inc.",
    "Republic Services Inc.", "Oneok, Inc.", "Ross Stores Inc", "Quanta Services, Inc.", "Cencora, Inc.",
    "Paychex Inc", "Becton, Dickinson and Co.", "Marathon Petroleum Corporation", "Paccar Inc",
    "Ameriprise Financial, Inc.", "General Motors Company", "American International Group, Inc.",
    "Sempra", "Simon Property Group, Inc.", "Realty Income Corporation", "Fedex Corporation",
    "W.W. Grainger, Inc.", "Te Connectivity Plc", "Dominion Energy, Inc", "Schlumberger Limited",
    "Corteva, Inc.", "Kimberly-Clark Corp.", "Fastenal Co", "Public Storage", "Phillips 66",
    "Kenvue Inc.", "The Kroger Co.", "Keurig Dr Pepper Inc.", "United Rentals, Inc.", "Metlife, Inc.",
    "Edwards Lifesciences Corp", "Exelon Corporation", "Cummins Inc.", "Crown Castle Inc.",
    "Verisk Analytics, Inc.", "Monster Beverage Corporation", "L3harris Technologies, Inc.", "Msci, Inc.",
    "Target Corporation", "Fidelity National Information Services, Inc.", "Fair Isaac Corporation",
    "Ametek, Inc.", "Idexx Laboratories Inc", "Yum! Brands, Inc.", "Ford Motor Company",
    "Valero Energy Corporation", "Xcel Energy, Inc.", "Charter Comm Inc Del Cl a",
    "Cognizant Technology Solutions", "Public Service Enterprise Group Incorporated",
    "Otis Worldwide Corporation", "Pg&e Corporation", "The Hartford Insurance Group, Inc.",
    "Corning Incorporated", "Take-Two Interactive Software Inc", "Cardinal Health, Inc.",
    "Consolidated Edison, Inc.", "Baker Hughes Company", "Resmed Inc.", "Prudential Financial, Inc.",
    "Hess Corporation", "Vulcan Materials Company", "Cbre Group, Inc.", "Electronic Arts Inc",
    "Lululemon Athletica Inc.", "Entergy Corporation", "Sysco Corporation", "Targa Resources Corp.",
    "Dell Technologies Inc.", "Arch Capital Group Ltd", "Wabtec Inc.", "Ebay Inc",
    "Martin Marietta Materials", "Rockwell Automation, Inc.", "Gartner, Inc.", "D.R. Horton Inc.",
    "Wec Energy Group, Inc.", "Nasdaq, Inc.", "Eqt Corp", "Dexcom, Inc.", "Vici Properties Inc.",
    "Ingersoll Rand Inc.", "Monolithic Power Systems, Inc.", "Equifax, Incorporated",
    "Ge Healthcare Technologies Inc.", "Microchip Technology Inc", "Nrg Energy, Inc.", "Costar Group Inc",
    "Agilent Technologies Inc.", "Willis Towers Watson Public Limited Companys", "Delta Air Lines, Inc.",
    "Old Dominion Freight Line", "Garmin Ltd", "Extra Space Storage, Inc.", "Xylem Inc",
    "Centene Corporation", "Ansys Inc", "M&t Bank Corp.", "Humana Inc.", "General Mills, Inc.",
    "Constellation Brands, Inc.", "Avalonbay Communities, Inc.", "Dupont De Nemours, Inc.",
    "Iron Mountain Inc.", "Dte Energy Company", "Keysight Technologies, Inc.",
    "Broadridge Financial Solutions Inc", "American Water Works Company, Inc", "Ventas, Inc.",
    "Occidental Petroleum Corporation", "State Street Corporation", "Raymond James Financial, Inc.",
    "Hp Inc.", "Tractor Supply Co", "Brown & Brown, Inc.", "Nucor Corporation", "Ameren Corporation",
    "Ppg Industries, Inc.", "Godaddy Inc", "Ppl Corporation", "International Paper Co.",
    "Texas Pacific Land Corporation", "Diamondback Energy, Inc.", "Fifth Third Bancorp",
    "Atmos Energy Corporation", "Veralto Corporation", "Sba Communications Corp",
    "Lennar Corporation Class A", "Dover Corporation", "Centerpoint Energy, Inc.",
    "Tyler Technologies, Inc.", "Cdw Corporation", "Steris Plc", "United Airlines Holdings, Inc.",
    "Expand Energy Corporation", "Fortive Corporation", "Cboe Global Markets, Inc.",
    "Mettler-Toledo International", "Darden Restaurants, Inc.", "Equity Residential", "Eversource Energy",
    "Church & Dwight Co., Inc.", "The Kraft Heinz Company", "Archer Daniels Midland Company",
    "Hewlett Packard Enterprise Company", "Iqvia Holdings Inc.", "Carnival Corporation", "Verisign Inc",
    "Cincinnati Financial Corp", "Corpay, Inc.", "Teledyne Technologies Incorporated", "Insulet Corporation",
    "Live Nation Entertainment Inc.", "The Hershey Company", "Synchrony Financial",
    "Huntington Bancshares Inc", "Dollar General Corp.", "Kellanova", "Edison International",
    "Seagate Technology Holdings Plcs", "Smurfit Westrock Plc", "W.R. Berkley Corporation",
    "Firstenergy Corp.", "Cms Energy Corporation", "Nvr, Inc.", "T Rowe Price Group Inc",
    "Super Micro Computer, Inc.", "Leidos Holdings, Inc.", "Amcor Plcs", "Devon Energy Corporation",
    "Williams-Sonoma, Inc.", "Northern Trust Corp", "Waters Corp", "Hubbell Incorporated", "Ptc, Inc",
    "Netapp, Inc", "Labcorp Holdings Inc.", "Dow Inc.", "Pultegroup, Inc.",
    "Warner Bros. Discovery, Inc. Series a", "International Flavors & Fragrances Inc.",
    "Expedia Group, Inc.", "Regions Financial Corp.", "Invitation Homes Inc.", "Quest Diagnostics Inc.",
    "Southwest Airlines Co.", "Ulta Beauty, Inc.", "Deckers Outdoor Corp", "Global Payments, Inc.",
    "Steel Dynamics Inc", "Biogen Inc.", "Zimmer Biomet Holdings, Inc.", "On Semiconductor Corp",
    "Weyerhaeuser Company", "Lennox International Inc.",
    "Mccormick & Company, Incorporated Non-Vtg Cs", "Mid-America Apartment Communities, Inc.",
    "Nisource Inc.", "Jabil Inc.", "Coterra Energy Inc.", "Dollar Tree Inc.", "Essex Property Trust, Inc",
    "Molina Healthcare, Inc.", "Factset Research Systems", "First Solar, Inc.", "Genuine Parts Company",
    "Citizens Financial Group, Inc.", "Halliburton Company", "Trimble Inc.", "Western Digital Corp.",
    "Packaging Corp of America", "Snap-on Incorporated", "Domino's Pizza Inc.", "F5, Inc.",
    "Tapestry, Inc.", "Principal Financial Group, Inc.", "Clorox Company", "Pentair Plc",
    "Ball Corporation", "Rollins, Inc.", "Tyson Foods, Inc.", "Alliant Energy Corporation",
    "Expeditors International of Washington, Inc.", "The Cooper Companies, Inc.",
    "Baxter International Inc.", "Cf Industries Holding, Inc.", "Loews Corporation", "Keycorp",
    "Jacobs Solutions Inc.", "Aptiv Plc", "Evergy, Inc.", "West Pharmaceutical Services, Inc.",
    "Gen Digital Inc.", "The Estee Lauder Companies Inc. Class A", "Zebra Technologies Corporation",
    "Lyondellbasell Industries N.v. Class A", "Everest Group, Ltd.", "Avery Dennison Corp.",
    "Omnicom Group Inc.", "Las Vegas Sands Corp.", "Kimco Realty Corp.", "Best Buy Company, Inc.",
    "Idex Corporation", "Textron, Inc.", "Henry (Jack) & Associates", "Masco Corporation",
    "Teradyne, Inc.", "Builders Firstsource, Inc.", "Camden Property Trust",
    "Allegion Public Limited Company", "Hologic Inc", "Udr, Inc.", "Paycom Software, Inc.",
    "Align Technology Inc", "Fox Corporation Class A", "The J.M. Smucker Company",
    "Juniper Networks Inc", "Healthpeak Properties, Inc.", "Regency Centers Corporation", "Pool Corporation",
    "Akamai Technologies Inc", "C.H. Robinson Worldwide, Inc.", "Skyworks Solutions Inc",
    "The Mosaic Company", "Universal Health Services, Inc. Class B", "Jb Hunt Transport Services Inc",
    "Revvity, Inc.", "Ralph Lauren Corporation", "Conagra Brands, Inc.", "Nordson Corp",
    "Tko Group Holdings, Inc.", "News Corporation Class A", "Alexandria Real Estate Equities, Inc.",
    "Lkq Corporation", "Bunge Global Sa", "Epam Systems, Inc.", "Incyte Genomics Inc",
    "Pinnacle West Capital Corporation", "Host Hotels & Resorts, Inc.", "Stanley Black & Decker, Inc.",
    "Viatris Inc.", "Globe Life Inc.", "Solventum Corporation", "Assurant, Inc.", "Carmax Inc.",
    "Molson Coors Beverage Company Class B", "Bxp, Inc.", "Eastman Chemical Company", "Dayforce, Inc.",
    "The Interpublic Group of Companies, Inc.", "Henry Schein Inc", "Erie Indemnity Co",
    "Huntington Ingalls Industries, Inc.", "Moderna, Inc.", "Hasbro, Inc.", "Hormel Foods Corporation",
    "Wynn Resorts Ltd", "Walgreens Boots Alliance, Inc", "A.O. Smith Corporation",
    "Norwegian Cruise Line Holdings Ltd.s", "Bio-Techne Corp.", "Match Group, Inc",
    "Lamb Weston Holdings, Inc.", "Generac Holdings Inc", "Aes Corporation",
    "Federal Realty Investment Trust", "Paramount Global Class B", "Mgm Resorts International",
    "Charles River Laboratories International, Inc.", "Fox Corporation Class B",
    "The Campbell's Company", "Franklin Resources, Inc.", "Albemarle Corporation", "Invesco Ltd",
    "Brown-Forman Corporation Class B", "Apa Corporation", "Davita Inc.", "Mohawk Industries, Inc.",
    "News Corporation Class B", "Amentum Holdings, Inc.", "AppLovin Corp.", "Robinhood Markets",
    "Emcor Group Inc."
]

SEC_TICKERS_URL = "https://www.sec.gov/files/company_tickers.json"

ALIASES = {
    "nextra energy": "NextEra Energy, Inc.",
    "marriot international": "Marriott International, Inc. (Class A)",
    "ge aerospace": "GE Aerospace",
    "ge vernova": "GE Vernova Inc.",
    "ge healthcare technologies": "GE HealthCare Technologies Inc.",
    "sandp global": "S&P Global Inc.",
    "s p global": "S&P Global Inc.",
    "abc amerisourcebergen": "Cencora, Inc.",
    "at t": "AT&T Inc.",
    "o reilly automotive": "O'Reilly Automotive, Inc.",
    "cvs health": "CVS Health Corporation",
    "cme group": "CME Group Inc.",
    "3m": "3M Company",
    "berkshire hathaway": "Berkshire Hathaway Inc.",
    "alphabet": "Alphabet Inc.",
    "jm smucker": "The J.M. Smucker Company",
    "campbells": "The Campbell Soup Company",
    "ge": "General Electric Company",
}

def _norm_title(s: str) -> str:
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = s.lower()
    s = re.sub(r"\bclass\s+[a-z0-9\-]+\b", "", s)
    s = s.replace("&", "and")
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    s = re.sub(r"\b(the|incorporated|inc|corp|corporation|co|company|plc|ltd|limited|reit|nv|sa)\b", "", s)
    return re.sub(r"\s+", " ", s).strip()

def fetch_json(url):
    r = requests.get(url, headers=HEADERS, timeout=30)
    r.raise_for_status()
    return r.json()

def build_target_companies_from_names(names):
    data = fetch_json(SEC_TICKERS_URL)
    rows = [{"ticker": v["ticker"].upper(), "cik": str(v["cik_str"]).zfill(10),
             "title": v["title"], "_norm": _norm_title(v["title"])} for v in data.values()]
    by_norm = {r["_norm"]: r for r in rows}
    
    out = []
    for name in names:
        raw_norm = _norm_title(name)
        alias_title = ALIASES.get(raw_norm)
        norm_key = _norm_title(alias_title) if alias_title else raw_norm
        r = by_norm.get(norm_key)
        if r:
            out.append((r["ticker"], r["cik"], r["title"]))
        else:
            print(f"[WARN] No match for: {name}")
    
    seen, dedup = set(), []
    for t in out:
        if t[0] not in seen:
            seen.add(t[0])
            dedup.append(t)
    return dedup

TARGET_COMPANIES = build_target_companies_from_names(REQUESTED_NAMES)
print(f"[INFO] Resolved {len(TARGET_COMPANIES)} companies")

# Comprehensive Tag groups
BALANCE_TAG_GROUPS = OrderedDict([
    ("Assets", [
        "Assets", "AssetsAbstract", "TotalAssets", "AssetsCurrent", "NoncurrentAssets",
        "AssetsNoncurrent", "PropertyPlantAndEquipmentNet", "Goodwill", "IntangibleAssetsNet"
    ]),
    ("AssetsCurrent", [
        "AssetsCurrent", "CurrentAssets", "CashAndCashEquivalentsAtCarryingValue",
        "AccountsReceivableNetCurrent", "InventoryNet", "PrepaidExpensesCurrent"
    ]),
    ("Liabilities", [
        "Liabilities", "LiabilitiesAbstract", "TotalLiabilities", "LiabilitiesCurrent",
        "LiabilitiesNoncurrent", "LongTermDebtNoncurrent", "AccountsPayableCurrent"
    ]),
    ("LiabilitiesCurrent", [
        "LiabilitiesCurrent", "CurrentLiabilities", "AccountsPayableCurrent",
        "AccruedLiabilitiesCurrent", "ShortTermDebt"
    ]),
    ("StockholdersEquity", [
        "StockholdersEquity", "StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest",
        "StockholdersEquityAbstract", "Equity", "CommonStockValue", "RetainedEarningsAccumulatedDeficit",
        "AdditionalPaidInCapital", "TreasuryStockValue"
    ]),
    ("LiabilitiesAndStockholdersEquity", [
        "LiabilitiesAndStockholdersEquity", "LiabilitiesAndStockholdersEquityAbstract",
        "LiabilitiesAndEquity", "TotalLiabilitiesAndStockholdersEquity"
    ]),
    ("CashAndCashEquivalents", [
        "CashAndCashEquivalentsAtCarryingValue", "CashAndCashEquivalents",
        "CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalents", "Cash",
        "CashAndDueFromBanks", "CashEquivalentsAtCarryingValue"
    ]),
    ("InventoryNet", ["InventoryNet", "Inventory", "InventoryGross", "MerchandiseInventory"]),
    ("LongTermDebt", [
        "LongTermDebtNoncurrent", "LongTermDebtAndCapitalLeaseObligations", "LongTermDebt",
        "DebtNoncurrent", "LongTermDebtExcludingCurrentMaturities"
    ]),
])

INCOME_TAG_GROUPS = OrderedDict([
    ("Revenues", [
        "Revenues", "SalesRevenueNet", "RevenueFromContractWithCustomerExcludingAssessedTax",
        "RevenueFromContractWithCustomer", "SalesRevenueGoodsNet", "SalesRevenueServicesNet",
        "TotalRevenue", "NetSales", "OperatingRevenues"
    ]),
    ("CostOfRevenue", [
        "CostOfRevenue", "CostOfGoodsAndServicesSold", "CostOfSales", "CostOfGoodsSold",
        "CostOfProductsSold", "CostOfServices"
    ]),
    ("GrossProfit", [
        "GrossProfit", "GrossMargin"
    ]),
    ("OperatingExpenses", [
        "OperatingExpenses", "OperatingExpensesGeneralAndAdministrative",
        "SellingGeneralAndAdministrativeExpenses", "ResearchAndDevelopmentExpense",
        "SellingAndMarketingExpense", "GeneralAndAdministrativeExpense", "SellingExpense",
        "AdministrativeExpense"
    ]),
    ("OperatingIncomeLoss", [
        "OperatingIncomeLoss",
        "IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest",
        "OperatingIncome", "IncomeFromOperations"
    ]),
    ("NetIncomeLoss", [
        "NetIncomeLoss", "ProfitLoss", "NetIncomeLossAvailableToCommonStockholdersBasic",
        "NetIncome", "ComprehensiveIncomeNetOfTax"
    ]),
    ("EarningsPerShareBasic", [
        "EarningsPerShareBasic", "EarningsPerShare"
    ]),
    ("EarningsPerShareDiluted", [
        "EarningsPerShareDiluted", "EarningsPerShareDilutedOne"
    ]),
])

INCOME_IS_CURRENCY = {
    "Revenues": True,
    "CostOfRevenue": True,
    "GrossProfit": True,
    "OperatingExpenses": True,
    "OperatingIncomeLoss": True,
    "NetIncomeLoss": True,
    "EarningsPerShareBasic": False,
    "EarningsPerShareDiluted": False,
}

# Data Extraction - ONLY REPORTED DATA
def get_company_facts(cik):
    return fetch_json(f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json")

def search_concept_deep(facts, concept_list, prefer_instant=False, select_form="latest"):
    """Search for concepts - returns ONLY actual reported values"""
    all_results = {}
    for concept in concept_list:
        try:
            fact = facts["facts"]["us-gaap"][concept]
        except KeyError:
            continue
        
        units = fact.get("units", {})
        unit = "USD" if "USD" in units else (next(iter(units)) if units else None)
        if not unit:
            continue
        
        obs = units[unit]
        if prefer_instant:
            obs = [o for o in obs if "instant" in o]
            date_key = "instant"
        else:
            obs = [o for o in obs if "start" in o and "end" in o]
            date_key = "end"
        
        if not obs:
            continue
        
        # Filter by form
        if select_form in ("10-K", "10-Q"):
            obs = [o for o in obs if o.get("form") == select_form]
        
        if not obs:
            continue
        
        # Group by year and take latest for each year
        for o in obs:
            year = o[date_key][:4]
            if START_YEAR <= int(year) <= END_YEAR:
                if year not in all_results:
                    all_results[year] = {}
                
                # Only take if we don't have data for this year or if this is more recent
                current_val = all_results[year].get("val")
                if current_val is None or o[date_key] > all_results[year]["date"]:
                    all_results[year] = {
                        "val": o.get("val"),
                        "date": o.get(date_key),
                        "unit": unit,
                        "form": o.get("form"),
                        "concept": concept
                    }
    
    return all_results

def build_balance_sheet(meta, facts, select_form):
    """Build balance sheet with ONLY reported data - no calculations"""
    print(f"  Building balance sheet for {meta['ticker']}...")
    yearly_data = {}
    
    for display, candidates in BALANCE_TAG_GROUPS.items():
        result = search_concept_deep(facts, candidates, prefer_instant=True, select_form=select_form)
        for year, data in result.items():
            if year not in yearly_data:
                yearly_data[year] = {
                    "Ticker": meta["ticker"],
                    "CIK": meta["cik"],
                    "Company": meta["title"],
                    "Year": year
                }
            # Only set if actually found in SEC data
            yearly_data[year][display] = data["val"]
    
    data_list = list(yearly_data.values())
    if data_list:
        years_found = [item['Year'] for item in data_list]
        print(f"  Found balance sheet data for years: {', '.join(years_found)}")
    return data_list

def build_income_statement(meta, facts, select_form):
    """Build income statement with reported data + calculated Gross Profit and Operating Income"""
    print(f"  Building income statement for {meta['ticker']}...")
    yearly_data = {}
    
    # First pass: collect all data except GrossProfit and OperatingIncomeLoss
    for display, candidates in INCOME_TAG_GROUPS.items():
        # Skip GrossProfit and OperatingIncomeLoss - we'll calculate them ourselves
        if display in ("GrossProfit", "OperatingIncomeLoss"):
            continue
            
        result = search_concept_deep(facts, candidates, prefer_instant=False, select_form=select_form)
        for year, data in result.items():
            if year not in yearly_data:
                yearly_data[year] = {
                    "Ticker": meta["ticker"],
                    "CIK": meta["cik"],
                    "Company": meta["title"],
                    "Year": year
                }
            # Only set if actually found in SEC data
            yearly_data[year][display] = data["val"]
            
            # Store revenues date
            if display == "Revenues":
                yearly_data[year]["Revenues (PeriodEnd)"] = data["date"]
    
    # Second pass: ALWAYS calculate Gross Profit and Operating Income for this specific company
    for year in list(yearly_data.keys()):
        data = yearly_data[year]
        revenues = data.get("Revenues")
        cost_of_revenue = data.get("CostOfRevenue")
        operating_expenses = data.get("OperatingExpenses")
        
        # Calculate Gross Profit
        if revenues is not None and cost_of_revenue is not None:
            # Calculate: Revenues - CostOfRevenue
            calculated_gp = revenues - cost_of_revenue
            yearly_data[year]["GrossProfit"] = calculated_gp
            print(f"    → Calculated Gross Profit for {year}: ${calculated_gp:,.0f}")
        else:
            # No cost data or no revenues, leave blank
            yearly_data[year]["GrossProfit"] = None
        
        # Calculate Operating Income/Loss
        gross_profit = yearly_data[year].get("GrossProfit")
        if gross_profit is not None and operating_expenses is not None:
            # Calculate: GrossProfit - OperatingExpenses
            calculated_oi = gross_profit - operating_expenses
            yearly_data[year]["OperatingIncomeLoss"] = calculated_oi
            print(f"    → Calculated Operating Income for {year}: ${calculated_oi:,.0f}")
        else:
            # Cannot calculate, set to None explicitly
            yearly_data[year]["OperatingIncomeLoss"] = None
    
    # Convert to list while maintaining year order
    data_list = [yearly_data[year] for year in sorted(yearly_data.keys())]
    
    if data_list:
        years_found = [item['Year'] for item in data_list]
        print(f"  Found income statement data for years: {', '.join(years_found)}")
    return data_list

def analyze_data_completeness(data_list, field_groups, data_type):
    """Analyze how complete the data is"""
    if not data_list:
        return 0, {}
    
    df = pd.DataFrame(data_list)
    total_possible = len(df) * len(field_groups)
    
    if total_possible == 0:
        return 0, {}
    
    completeness_by_field = {}
    for field in field_groups:
        non_null_count = df[field].notna().sum()
        completeness_by_field[field] = non_null_count / len(df) * 100
    
    overall_completeness = df[list(field_groups.keys())].notna().sum().sum() / total_possible * 100
    
    return overall_completeness, completeness_by_field

# Formatting
def format_currency_sheet(writer, sheet_name, currency_cols):
    ws = writer.sheets[sheet_name]
    header_to_idx = {cell.value: i + 1 for i, cell in enumerate(ws[1])}
    fmt = "$#,##0"
    for col_name in currency_cols:
        if col_name not in header_to_idx:
            continue
        col_letter = get_column_letter(header_to_idx[col_name])
        for r in range(2, ws.max_row + 1):
            cell = ws[f"{col_letter}{r}"]
            if isinstance(cell.value, (int, float)):
                cell.number_format = fmt

def create_yearly_sheets(writer, df_bs, df_is):
    """Create separate sheets for each year from 2020 to 2025 with actual data"""
    available_years_bs = sorted(df_bs["Year"].unique()) if not df_bs.empty else []
    available_years_is = sorted(df_is["Year"].unique()) if not df_is.empty else []
    all_available_years = sorted(set(available_years_bs + available_years_is))
    
    print(f"[INFO] Available years with data: {all_available_years}")
    
    for year in range(START_YEAR, END_YEAR + 1):
        year_str = str(year)
        
        # Balance sheet for the year
        bs_year_data = df_bs[df_bs["Year"] == year_str].drop(columns=["Year"]) if not df_bs.empty else pd.DataFrame()
        if not bs_year_data.empty:
            bs_year_data.to_excel(writer, sheet_name=f"BS_{year}", index=False)
            format_currency_sheet(writer, f"BS_{year}", list(BALANCE_TAG_GROUPS.keys()))
            print(f"[INFO] Created BS_{year} with {len(bs_year_data)} companies")
        
        # Income statement for the year
        is_year_data = df_is[df_is["Year"] == year_str].drop(columns=["Year"]) if not df_is.empty else pd.DataFrame()
        if not is_year_data.empty:
            is_year_data.to_excel(writer, sheet_name=f"IS_{year}", index=False)
            currency_is_cols = [c for c in list(INCOME_TAG_GROUPS.keys()) if INCOME_IS_CURRENCY.get(c)]
            format_currency_sheet(writer, f"IS_{year}", currency_is_cols)
            print(f"[INFO] Created IS_{year} with {len(is_year_data)} companies")

# ===================== Main =====================
def main():
    universe = [{"ticker": t, "cik": cik, "title": name} for (t, cik, name) in TARGET_COMPANIES]
    print(f"[INFO] Processing {len(universe)} companies...")
    print(f"[INFO] Extracting data for years {START_YEAR} to {END_YEAR}")
    print(f"[INFO] ONLY extracting reported data - NO calculations except Gross Profit")
    
    all_bs_data, all_is_data = [], []
    
    for i, meta in enumerate(universe):
        t, cik = meta["ticker"], meta["cik"]
        print(f"\n[{i+1}/{len(universe)}] Processing {t} ({meta['title']})")
        
        try:
            facts = get_company_facts(cik)
            
            # Get balance sheet data - ONLY reported values
            bs_data = build_balance_sheet(meta, facts, SELECT_FORM)
            if bs_data:
                all_bs_data.extend(bs_data)
            else:
                print(f"  No balance sheet data found")
            
            # Get income statement data - ONLY reported values + calculated GP
            is_data = build_income_statement(meta, facts, SELECT_FORM)
            if is_data:
                all_is_data.extend(is_data)
            else:
                print(f"  No income statement data found")
                
        except Exception as e:
            print(f"[ERROR] {t}: {e}")
        
        time.sleep(PAUSE_SECONDS)
    
    # Create DataFrames
    df_bs = pd.DataFrame(all_bs_data)
    df_is = pd.DataFrame(all_is_data)
    
    # Analyze overall completeness
    if not df_bs.empty:
        bs_completeness, bs_field_stats = analyze_data_completeness(all_bs_data, BALANCE_TAG_GROUPS, "Balance Sheet")
        print(f"\n[INFO] Balance Sheet Overall Completeness: {bs_completeness:.1f}%")
        print(f"[INFO] Balance Sheet Field Completeness:")
        for field, completeness in bs_field_stats.items():
            print(f"  {field}: {completeness:.1f}%")
    
    if not df_is.empty:
        is_completeness, is_field_stats = analyze_data_completeness(all_is_data, INCOME_TAG_GROUPS, "Income Statement")
        print(f"\n[INFO] Income Statement Overall Completeness: {is_completeness:.1f}%")
        print(f"[INFO] Income Statement Field Completeness:")
        for field, completeness in is_field_stats.items():
            print(f"  {field}: {completeness:.1f}%")
    
    # Reorder columns for income statement
    if not df_is.empty and "Revenues (PeriodEnd)" in df_is.columns:
        base_cols = ["Ticker", "CIK", "Company", "Year"]
        revenue_cols = ["Revenues", "Revenues (PeriodEnd)"]
        other_is_cols = [c for c in df_is.columns if c not in base_cols + revenue_cols and not c.endswith("(PeriodEnd)")]
        df_is = df_is[base_cols + revenue_cols + other_is_cols]
    
    # Create Excel file
    with pd.ExcelWriter(OUT_FILE, engine="openpyxl") as writer:
        # Consolidated sheets
        if not df_bs.empty:
            df_bs.to_excel(writer, sheet_name="BalanceSheet_Consolidated", index=False)
            format_currency_sheet(writer, "BalanceSheet_Consolidated", list(BALANCE_TAG_GROUPS.keys()))
            print(f"[INFO] Created BalanceSheet_Consolidated with {len(df_bs)} records")
        
        if not df_is.empty:
            df_is.to_excel(writer, sheet_name="IncomeStatement_Consolidated", index=False)
            currency_is_cols = [c for c in list(INCOME_TAG_GROUPS.keys()) if INCOME_IS_CURRENCY.get(c)]
            format_currency_sheet(writer, "IncomeStatement_Consolidated", currency_is_cols)
            print(f"[INFO] Created IncomeStatement_Consolidated with {len(df_is)} records")
        
        # Yearly sheets
        print(f"\n[INFO] Creating yearly sheets...")
        create_yearly_sheets(writer, df_bs, df_is)
    
    print(f"\n✓ Saved {OUT_FILE}")
    print(f"✓ Extraction completed successfully!")
    
    if not df_bs.empty or not df_is.empty:
        all_years = sorted(set(df_bs['Year'].unique()) | set(df_is['Year'].unique()))
        print(f"✓ Data extracted for years: {all_years}")
        print(f"✓ Total companies processed: {len(universe)}")
        companies_with_data = len(set(df_bs['Ticker'].unique()) | set(df_is['Ticker'].unique()))
        print(f"✓ Companies with data: {companies_with_data}")
        print(f"✓ Overall data coverage: {companies_with_data/len(universe)*100:.1f}%")
    else:
        print("⚠ No data was extracted. Please check the company names and SEC data availability.")

if __name__ == "__main__":
    main()
